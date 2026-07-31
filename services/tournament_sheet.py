"""Extract historical tournament standings and pairings from the shared workbook."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pydantic import BaseModel, Field


class SheetStanding(BaseModel):
    place: int = Field(ge=1)
    player: str = Field(min_length=1)
    deck: str = Field(min_length=1)


class SheetPairing(BaseModel):
    player1: str = Field(min_length=1)
    player2: str = Field(min_length=1)
    result1: int = Field(ge=0)
    result2: int = Field(ge=0)


class SheetRound(BaseModel):
    number: int = Field(ge=1)
    pairings: list[SheetPairing] = Field(min_length=1)


class SheetTournament(BaseModel):
    date: date
    club: str = Field(min_length=1)
    format: str = "Pauper"
    standings: list[SheetStanding] = Field(min_length=1)
    rounds: list[SheetRound] = Field(min_length=1)


class TournamentSheetError(ValueError):
    pass


def _event_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError:
            return None
    return None


def _text(value: object) -> str:
    return str(value or "").strip()


def _headers(row: tuple[object, ...]) -> dict[str, int]:
    return {_text(value).casefold(): index for index, value in enumerate(row) if _text(value)}


def _rounds(pairings: list[SheetPairing]) -> list[SheetRound]:
    rounds: list[SheetRound] = []
    current: list[SheetPairing] = []
    seen: set[str] = set()
    for pairing in pairings:
        names = {pairing.player1.casefold(), pairing.player2.casefold()}
        if current and seen.intersection(names):
            rounds.append(SheetRound(number=len(rounds) + 1, pairings=current))
            current = []
            seen = set()
        current.append(pairing)
        seen.update(names)
    if current:
        rounds.append(SheetRound(number=len(rounds) + 1, pairings=current))
    return rounds


def extract_sheet_tournaments(
    workbook_path: str | Path,
    target_keys: set[tuple[str, date]],
) -> tuple[list[SheetTournament], list[str]]:
    """Return requested tournaments and non-fatal data issues.

    ``Match History`` stores every match twice, once from each player's point of
    view. Reciprocal adjacent rows are collapsed before round boundaries are
    inferred from the first repeated player.
    """

    normalized_targets = {(club.casefold(), event_date): club for club, event_date in target_keys}
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Pauper" not in workbook.sheetnames or "Match History" not in workbook.sheetnames:
        raise TournamentSheetError('Ожидались листы "Pauper" и "Match History"')

    standings_by_key: dict[tuple[str, date], list[SheetStanding]] = {}
    standings_sheet = workbook["Pauper"]
    rows = standings_sheet.iter_rows(values_only=True)
    standings_headers = _headers(next(rows))
    required = {"место", "участник", "колода", "клуб", "турниры"}
    if not required.issubset(standings_headers):
        raise TournamentSheetError(f"В листе Pauper отсутствуют колонки: {sorted(required - standings_headers.keys())}")
    for row in rows:
        event_date = _event_date(row[standings_headers["турниры"]])
        club = _text(row[standings_headers["клуб"]])
        key = (club.casefold(), event_date) if event_date else None
        if key not in normalized_targets:
            continue
        try:
            standing = SheetStanding(
                place=int(row[standings_headers["место"]]),
                player=_text(row[standings_headers["участник"]]),
                deck=_text(row[standings_headers["колода"]]),
            )
        except (TypeError, ValueError) as exc:
            raise TournamentSheetError(f"Некорректный standing для {club} {event_date}: {exc}") from exc
        standings_by_key.setdefault(key, []).append(standing)

    history_by_key: dict[tuple[str, date], list[tuple[str, str, int, int]]] = {}
    history_sheet = workbook["Match History"]
    rows = history_sheet.iter_rows(values_only=True)
    history_headers = _headers(next(rows))
    required = {"дата", "игрок", "оппонент", "счет игрока", "счет оппонента", "клуб"}
    if not required.issubset(history_headers):
        raise TournamentSheetError(
            f"В листе Match History отсутствуют колонки: {sorted(required - history_headers.keys())}"
        )
    for row in rows:
        event_date = _event_date(row[history_headers["дата"]])
        club = _text(row[history_headers["клуб"]])
        key = (club.casefold(), event_date) if event_date else None
        if key not in normalized_targets:
            continue
        try:
            history_by_key.setdefault(key, []).append(
                (
                    _text(row[history_headers["игрок"]]),
                    _text(row[history_headers["оппонент"]]),
                    int(row[history_headers["счет игрока"]]),
                    int(row[history_headers["счет оппонента"]]),
                )
            )
        except (TypeError, ValueError) as exc:
            raise TournamentSheetError(f"Некорректный матч для {club} {event_date}: {exc}") from exc

    issues: list[str] = []
    tournaments: list[SheetTournament] = []
    for (club_key, event_date), display_club in sorted(
        normalized_targets.items(), key=lambda item: (item[0][1], item[0][0])
    ):
        standings = sorted(standings_by_key.get((club_key, event_date), []), key=lambda row: row.place)
        directed = history_by_key.get((club_key, event_date), [])
        if not standings or not directed:
            issues.append(f"{event_date.isoformat()} {club_key}: нет standings или pairings в таблице")
            continue
        pairings: list[SheetPairing] = []
        index = 0
        while index < len(directed):
            player1, player2, result1, result2 = directed[index]
            reciprocal = index + 1 < len(directed) and directed[index + 1] == (player2, player1, result2, result1)
            if not reciprocal:
                issues.append(
                    f"{event_date.isoformat()} {club_key}: у матча {player1} — {player2} нет соседней обратной строки"
                )
                index += 1
                continue
            pairings.append(SheetPairing(player1=player1, player2=player2, result1=result1, result2=result2))
            index += 2
        if not pairings:
            issues.append(f"{event_date.isoformat()} {club_key}: после дедупликации не осталось pairings")
            continue
        expected_places = list(range(1, len(standings) + 1))
        actual_places = [row.place for row in standings]
        if actual_places != expected_places:
            issues.append(f"{event_date.isoformat()} {club_key}: места {actual_places}, ожидались {expected_places}")
            continue
        tournaments.append(
            SheetTournament(
                date=event_date,
                club=display_club,
                standings=standings,
                rounds=_rounds(pairings),
            )
        )
    return tournaments, issues
