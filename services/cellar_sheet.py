"""Public Google Sheets source for the Edinorog cellar catalog."""

from __future__ import annotations

import hashlib
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import BinaryIO
from urllib.parse import urlparse

import requests
from openpyxl import load_workbook

CELLAR_SHEET_ID = "1SJbffwDLqSFBg4assNJ0zPUIpObrIcY6VuY1otgt4lI"
CELLAR_SHEET_URL = f"https://docs.google.com/spreadsheets/d/{CELLAR_SHEET_ID}/edit"
CELLAR_SHEET_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{CELLAR_SHEET_ID}/export?format=xlsx"
CELLAR_SHEET_HEADER = "колода \\ дата"
CELLAR_SHEET_MAX_BYTES = 10 * 1024 * 1024
CELLAR_UNAVAILABLE_NOTES = {"дома", "занято"}


class CellarCatalogSourceError(RuntimeError):
    pass


@dataclass(frozen=True)
class CatalogEntry:
    source_key: str
    name: str
    archetype_name: str
    decklist_url: str | None = None
    notes: str | None = None
    decklist_updated_on: date | None = None
    available: bool = True


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = " ".join(str(value).split())
    return text or None


def _source_key(name: str, copy_number: int) -> str:
    normalized = name.casefold().encode("utf-8")
    digest = hashlib.sha256(normalized).hexdigest()[:16]
    return f"gsheet:{digest}:{copy_number}"


def _safe_http_url(value: str | None) -> str | None:
    if value is None:
        return None
    parsed = urlparse(value)
    return value if parsed.scheme in {"http", "https"} and parsed.netloc else None


def parse_cellar_workbook(workbook: BinaryIO | BytesIO) -> list[CatalogEntry]:
    """Read deck rows B:D, including hyperlinks attached to column B."""

    try:
        sheet = load_workbook(workbook, data_only=True, read_only=False).active
    except Exception as exc:
        raise CellarCatalogSourceError("Не удалось прочитать XLSX-каталог ячейки.") from exc

    header_row = next(
        (
            row
            for row in range(1, sheet.max_row + 1)
            if (_clean_text(sheet.cell(row, 2).value) or "").casefold() == CELLAR_SHEET_HEADER
        ),
        None,
    )
    if header_row is None:
        raise CellarCatalogSourceError("В таблице ячейки не найден столбец «Колода \\ Дата».")

    copies: defaultdict[str, int] = defaultdict(int)
    entries: list[CatalogEntry] = []
    for row in range(header_row + 1, sheet.max_row + 1):
        name = _clean_text(sheet.cell(row, 2).value)
        if name is None:
            continue
        normalized_name = name.casefold()
        copies[normalized_name] += 1

        notes = _clean_text(sheet.cell(row, 3).value)
        updated_value = sheet.cell(row, 4).value
        if isinstance(updated_value, datetime):
            updated_on = updated_value.date()
        elif isinstance(updated_value, date):
            updated_on = updated_value
        else:
            updated_on = None
        hyperlink = sheet.cell(row, 2).hyperlink
        decklist_url = _safe_http_url(hyperlink.target if hyperlink is not None else None)
        entries.append(
            CatalogEntry(
                source_key=_source_key(name, copies[normalized_name]),
                name=name,
                archetype_name=name,
                decklist_url=decklist_url,
                notes=notes,
                decklist_updated_on=updated_on,
                available=(notes or "").casefold() not in CELLAR_UNAVAILABLE_NOTES,
            )
        )

    if not entries:
        raise CellarCatalogSourceError("В таблице ячейки не найдено ни одной колоды.")
    return entries


class GoogleSheetsCellarCatalog:
    def __init__(self, *, session: requests.Session | None = None, timeout: int = 10) -> None:
        self._session = session or requests.Session()
        self._timeout = timeout

    def fetch(self) -> list[CatalogEntry]:
        try:
            response = self._session.get(CELLAR_SHEET_EXPORT_URL, timeout=self._timeout)
            response.raise_for_status()
        except requests.RequestException as exc:
            raise CellarCatalogSourceError("Не удалось загрузить каталог ячейки из Google Sheets.") from exc

        if len(response.content) > CELLAR_SHEET_MAX_BYTES:
            raise CellarCatalogSourceError("XLSX-каталог ячейки превышает допустимый размер.")
        return parse_cellar_workbook(BytesIO(response.content))
