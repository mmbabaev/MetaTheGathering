from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import List, Literal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import nulls_last, select
from sqlalchemy.orm import Session

from bot.messages import format_participant_name
from core import models
from services.stats import StatsService
from services.utils import get_tournament

ExportFormat = Literal["csv", "markdown"]


class ExportService:
    def __init__(self, db: Session):
        self.db = db

    def _participants_query(self, tournament_id: int):
        return (
            select(
                models.Participant.id,
                models.User.username,
                models.User.tg_id,
                models.Archetype.name.label("archetype_name"),
                models.Participant.upvotes_count,
                models.Participant.downvotes_count,
                models.Participant.confirmed,
                models.Participant.added_by_admin,
                models.Participant.created_at,
            )
            .join(models.User, models.User.id == models.Participant.user_id)
            .join(
                models.Archetype,
                models.Archetype.id == models.Participant.archetype_id,
                isouter=True,
            )
            .where(models.Participant.tournament_id == tournament_id)
            .order_by(nulls_last(models.Participant.final_place.asc()), models.Participant.created_at.asc())
        )

    def export_participants_csv(
        self,
        tournament_id: int,
        file: io.TextIOBase | None = None,
        encoding: str = "utf-8",
    ) -> str:
        """
        Выгрузка участников турнира в CSV.
        Если file None — возвращает строку, иначе пишет в указанный файл и возвращает пустую строку.
        """
        get_tournament(self.db, tournament_id)  # ensure exists

        stmt = self._participants_query(tournament_id)
        rows = self.db.execute(stmt).all()

        headers = [
            "participant_id",
            "username",
            "tg_id",
            "archetype",
            "upvotes",
            "downvotes",
            "confirmed",
            "added_by_admin",
            "created_at",
        ]

        if file is not None:
            writer = csv.writer(file)
            writer.writerow(headers)
            for r in rows:
                writer.writerow(
                    [
                        r.id,
                        r.username or "",
                        r.tg_id,
                        r.archetype_name or "",
                        r.upvotes_count,
                        r.downvotes_count,
                        int(r.confirmed),
                        int(r.added_by_admin),
                        r.created_at.isoformat() if isinstance(r.created_at, datetime) else "",
                    ]
                )
            return ""

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        for r in rows:
            writer.writerow(
                [
                    r.id,
                    r.username or "",
                    r.tg_id,
                    r.archetype_name or "",
                    r.upvotes_count,
                    r.downvotes_count,
                    int(r.confirmed),
                    int(r.added_by_admin),
                    r.created_at.isoformat() if isinstance(r.created_at, datetime) else "",
                ]
            )
        return buf.getvalue()

    def export_players_list(self, tournament_id: int) -> str:
        """Возвращает plain-text «Имя Фамилия» по одному на строку, отсортировано по финальным стендингам."""
        participants = self.db.query(models.Participant).filter_by(tournament_id=tournament_id).all()
        participants.sort(
            key=lambda p: (
                p.final_place if p.final_place is not None else 999999,
                format_participant_name(
                    p.user.first_name if p.user else None,
                    p.user.last_name if p.user else None,
                ).lower(),
            )
        )
        names = [
            format_participant_name(
                p.user.first_name if p.user else None,
                p.user.last_name if p.user else None,
            )
            for p in participants
        ]
        return "\n".join(names)

    def export_participants_excel(self, tournament_id: int) -> tuple[bytes, str]:
        """Возвращает (bytes, filename) для Excel-файла списка участников."""
        t = get_tournament(self.db, tournament_id)
        participants = self.db.query(models.Participant).filter_by(tournament_id=tournament_id).all()
        participants.sort(
            key=lambda p: (
                p.final_place if p.final_place is not None else 999999,
                format_participant_name(
                    p.user.first_name if p.user else None,
                    p.user.last_name if p.user else None,
                ).lower(),
            )
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Участники"

        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        headers = ["#", "@Ник", "Имя Фамилия"] if t.decks_hidden else ["#", "@Ник", "Имя Фамилия", "Колода"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, p in enumerate(participants, 2):
            username = f"@{p.user.username}" if p.user and p.user.username else ""
            full_name = format_participant_name(
                p.user.first_name if p.user else None,
                p.user.last_name if p.user else None,
            )
            ws.cell(row=row, column=1, value=p.final_place)
            ws.cell(row=row, column=2, value=username)
            ws.cell(row=row, column=3, value=full_name)
            if not t.decks_hidden:
                deck = p.archetype.name if p.archetype else ""
                ws.cell(row=row, column=4, value=deck)

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 28
        ws.column_dimensions["D"].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"{t.title.replace(' ', '_')}.xlsx"
        return buf.getvalue(), filename

    def get_pairings_by_round(self, tournament_id: int) -> list[tuple[int, list[tuple[str, str, object, object, str]]]]:
        """Паринги, сгруппированные по раундам.

        ``[(round_number, [(date, player1, result1, result2, player2), …]), …]``.
        ``date`` — дата турнира (dd.mm.yyyy) или "". ``result1``/``result2`` — победы
        в партиях каждого игрока, если известны (иначе ""). Баи пропускаются.
        """
        t = get_tournament(self.db, tournament_id)
        date_str = t.started_at.strftime("%d.%m.%Y") if t.started_at else ""

        def _score(value) -> object:
            return value if value is not None else ""

        by_round: dict[int, list[models.RoundPairing]] = {}
        for p in self.db.query(models.RoundPairing).filter_by(tournament_id=tournament_id).all():
            by_round.setdefault(p.round_number, []).append(p)

        groups: list[tuple[int, list[tuple[str, str, object, object, str]]]] = []
        for rnd in sorted(by_round):
            pairings = sorted(
                by_round[rnd],
                key=lambda p: (p.table_number if p.table_number is not None else 10**9, p.player_name),
            )
            rows: list[tuple[str, str, object, object, str]] = []
            seen: set[frozenset[str]] = set()
            for p in pairings:
                if not p.opponent_name:  # bye — в таблицу матчей не идёт
                    continue
                key = frozenset((p.player_name, p.opponent_name))
                if key in seen:
                    continue
                seen.add(key)
                rows.append((date_str, p.player_name, _score(p.player_wins), _score(p.opponent_wins), p.opponent_name))
            if rows:
                groups.append((rnd, rows))
        return groups

    def get_pairings_rows(self, tournament_id: int) -> list[tuple[str, str, object, object, str]]:
        """Плоский список матчей в «формате Серёжи» (одна строка на матч)."""
        return [row for _, rows in self.get_pairings_by_round(tournament_id) for row in rows]

    def export_pairings_excel(self, tournament_id: int) -> tuple[bytes, str] | None:
        """Excel с парингами в «формате Серёжи», с секциями по раундам. None, если парингов нет."""
        t = get_tournament(self.db, tournament_id)
        groups = self.get_pairings_by_round(tournament_id)
        if not groups:
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pairings"

        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        section_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        section_font = Font(bold=True)
        headers = ["date", "player1", "result1", "result2", "player2"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        row = 2
        for round_number, rows in groups:
            # секция раунда: заголовок на всю ширину
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = section_fill
            head = ws.cell(row=row, column=1, value=f"Раунд {round_number}")
            head.font = section_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
            for date_str, player1, result1, result2, player2 in rows:
                ws.cell(row=row, column=1, value=date_str)
                ws.cell(row=row, column=2, value=player1)
                ws.cell(row=row, column=3, value=result1)
                ws.cell(row=row, column=4, value=result2)
                ws.cell(row=row, column=5, value=player2)
                row += 1

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 26
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 9
        ws.column_dimensions["E"].width = 26

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"{t.title.replace(' ', '_')}_pairings.xlsx"
        return buf.getvalue(), filename

    def export_meta_markdown(self, tournament_id: int) -> str:
        """
        Markdown‑таблица метагейма: архетип, количество игроков, суммарные голоса.
        Удобно кидать в Telegram / на сайт как текст.[web:61][web:67]
        """
        stats = StatsService(self.db)
        meta = stats.get_tournament_meta(tournament_id)

        if not meta:
            return "| Archetype | Players | Upvotes | Downvotes |\n|---|---|---|---|\n"

        lines: List[str] = []
        lines.append("| Archetype | Players | Upvotes | Downvotes |")
        lines.append("|---|---|---|---|")
        for row in meta:
            lines.append(f"| {row.archetype_name} | {row.count} | {row.upvotes_sum} | {row.downvotes_sum} |")
        return "\n".join(lines)
