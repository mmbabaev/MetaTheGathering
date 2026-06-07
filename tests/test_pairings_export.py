"""Tests for the Серёжа-format pairings export (ExportService)."""

import io
from datetime import datetime

import openpyxl

from core import models
from core.schemas import TournamentCreate
from services.export import ExportService
from services.tournament import TournamentService


def _tournament(db, title="Pauper Friday", chat_id=100, started_at=None):
    t = TournamentService(db).create_tournament(TournamentCreate(title=title, chat_id=chat_id))
    if started_at is not None:
        obj = db.get(models.Tournament, t.id)
        obj.started_at = started_at
        db.commit()
    return t


def _pairing(db, tid, rnd, p1, p2, table=None):
    db.add(
        models.RoundPairing(tournament_id=tid, round_number=rnd, player_name=p1, opponent_name=p2, table_number=table)
    )
    db.commit()


def _both(db, tid, rnd, a, b, table):
    # AetherHub/mtgarena store both directions of a match
    _pairing(db, tid, rnd, a, b, table)
    _pairing(db, tid, rnd, b, a, table)


def test_no_pairings_returns_none(db):
    t = _tournament(db)
    assert ExportService(db).export_pairings_excel(t.id) is None
    assert ExportService(db).get_pairings_rows(t.id) == []


def test_rows_one_per_match_deduped(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25, 19, 0))
    _both(db, t.id, 1, "Харитонов Алексей", "Давыдов Олег", 1)
    _both(db, t.id, 1, "Старостин Владислав", "Карпачев Денис", 2)
    rows = ExportService(db).get_pairings_rows(t.id)
    assert len(rows) == 2  # both directions collapsed to one row each
    dates = {r[0] for r in rows}
    assert dates == {"25.11.2024"}


def test_rows_skip_byes(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    _both(db, t.id, 1, "A", "B", 1)
    _pairing(db, t.id, 1, "C", None, None)  # bye
    rows = ExportService(db).get_pairings_rows(t.id)
    matches = {(r[1], r[4]) for r in rows}
    assert matches == {("A", "B")}
    assert all("C" not in (r[1], r[4]) for r in rows)


def test_results_blank_when_unknown(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    _both(db, t.id, 1, "A", "B", 1)
    (_, _, result1, result2, _) = ExportService(db).get_pairings_rows(t.id)[0]
    assert result1 == "" and result2 == ""


def test_results_filled_when_known(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    db.add(
        models.RoundPairing(
            tournament_id=t.id,
            round_number=1,
            player_name="A",
            opponent_name="B",
            table_number=1,
            player_wins=2,
            opponent_wins=1,
        )
    )
    db.add(
        models.RoundPairing(
            tournament_id=t.id,
            round_number=1,
            player_name="B",
            opponent_name="A",
            table_number=1,
            player_wins=1,
            opponent_wins=2,
        )
    )
    db.commit()
    rows = ExportService(db).get_pairings_rows(t.id)
    assert len(rows) == 1
    _, p1, r1, r2, p2 = rows[0]
    assert (p1, r1, r2, p2) == ("A", 2, 1, "B")


def test_no_started_at_blank_date(db):
    t = _tournament(db)  # no started_at
    _both(db, t.id, 1, "A", "B", 1)
    assert ExportService(db).get_pairings_rows(t.id)[0][0] == ""


def test_excel_headers_and_content(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    _both(db, t.id, 1, "Харитонов Алексей", "Давыдов Олег", 1)
    result = ExportService(db).export_pairings_excel(t.id)
    assert result is not None
    data, filename = result
    assert filename.endswith("_pairings.xlsx")

    ws = openpyxl.load_workbook(io.BytesIO(data)).active
    assert [c.value for c in ws[1]] == ["date", "player1", "result1", "result2", "player2"]
    assert ws.cell(row=2, column=1).value == "Раунд 1"  # секция раунда
    row3 = [c.value for c in ws[3]]
    assert row3[0] == "25.11.2024"
    assert {row3[1], row3[4]} == {"Харитонов Алексей", "Давыдов Олег"}


def test_excel_round_sections(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    _both(db, t.id, 1, "A", "B", 1)
    _both(db, t.id, 2, "A", "B", 1)
    ws = openpyxl.load_workbook(io.BytesIO(ExportService(db).export_pairings_excel(t.id)[0])).active
    section_headers = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if str(ws.cell(row=r, column=1).value or "").startswith("Раунд")
    ]
    assert section_headers == ["Раунд 1", "Раунд 2"]


def test_get_pairings_by_round(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    _both(db, t.id, 2, "E", "F", 1)
    _both(db, t.id, 1, "A", "B", 1)
    groups = ExportService(db).get_pairings_by_round(t.id)
    assert [rnd for rnd, _ in groups] == [1, 2]
    assert [(r[1], r[4]) for r in groups[0][1]] == [("A", "B")]


def test_ordered_by_round_then_table(db):
    t = _tournament(db, started_at=datetime(2024, 11, 25))
    _both(db, t.id, 2, "E", "F", 1)
    _both(db, t.id, 1, "C", "D", 2)
    _both(db, t.id, 1, "A", "B", 1)
    rows = ExportService(db).get_pairings_rows(t.id)
    # round 1 (table 1, then table 2), then round 2
    matchups = [(r[1], r[4]) for r in rows]
    assert matchups == [("A", "B"), ("C", "D"), ("E", "F")]
