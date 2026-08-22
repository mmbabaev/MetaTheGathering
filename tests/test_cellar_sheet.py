from datetime import date, datetime, timedelta
from io import BytesIO

import pytest
import requests
from openpyxl import Workbook

from core import models
from services.cellar import CELLAR_CATALOG_REFRESH_INTERVAL, CellarDeckUnavailable, CellarService
from services.cellar_sheet import (
    CatalogEntry,
    CellarCatalogSourceError,
    GoogleSheetsCellarCatalog,
    parse_cellar_workbook,
)


def _workbook_bytes(rows: list[tuple[object, object, object, str | None]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet["B11"] = "Колода \\ Дата"
    sheet["C11"] = "Доп. инфо"
    sheet["D11"] = "Актуальность листа"
    for row_number, (name, notes, updated_on, url) in enumerate(rows, start=12):
        sheet.cell(row_number, 2, name)
        sheet.cell(row_number, 3, notes)
        sheet.cell(row_number, 4, updated_on)
        if url:
            sheet.cell(row_number, 2).hyperlink = url
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def test_parse_sheet_preserves_physical_copies_links_dates_and_availability():
    entries = parse_cellar_workbook(
        BytesIO(
            _workbook_bytes(
                [
                    ("Altar Tron", None, datetime(2026, 7, 1), "https://example.test/altar"),
                    ("Altar Tron", "занято", None, None),
                    ("Dimir Control", "дома", None, "https://example.test/dimir"),
                    ("Unsafe link", None, None, "javascript:alert(1)"),
                ]
            )
        )
    )

    assert [entry.name for entry in entries] == ["Altar Tron", "Altar Tron", "Dimir Control", "Unsafe link"]
    assert len({entry.source_key for entry in entries}) == 4
    assert entries[0].source_key.removesuffix(":1") == entries[1].source_key.removesuffix(":2")
    assert entries[0].decklist_url == "https://example.test/altar"
    assert entries[0].decklist_updated_on == date(2026, 7, 1)
    assert entries[0].available is True
    assert entries[1].available is False
    assert entries[2].available is False
    assert entries[3].decklist_url is None


def test_parse_sheet_rejects_unknown_layout():
    workbook = Workbook()
    content = BytesIO()
    workbook.save(content)

    with pytest.raises(CellarCatalogSourceError, match="Колода"):
        parse_cellar_workbook(BytesIO(content.getvalue()))


def test_sheet_client_wraps_network_errors():
    class BrokenSession:
        def get(self, *_args, **_kwargs):
            raise requests.ConnectionError("offline")

    with pytest.raises(CellarCatalogSourceError, match="Google Sheets"):
        GoogleSheetsCellarCatalog(session=BrokenSession()).fetch()


def test_catalog_sync_updates_rows_and_deactivates_removed_decks(db):
    service = CellarService(db)
    old = CatalogEntry("gsheet:old:1", "Old name", "Old name")
    removed = CatalogEntry("gsheet:removed:1", "Removed", "Removed")
    service.sync_catalog([old, removed], synced_at=datetime(2026, 8, 22, 10, 0))

    result = service.sync_catalog(
        [
            CatalogEntry(
                "gsheet:old:1",
                "New name",
                "New archetype",
                decklist_url="https://example.test/new",
                notes="дома",
                decklist_updated_on=date(2026, 7, 1),
                available=False,
            ),
            CatalogEntry("gsheet:new:1", "New deck", "New deck"),
        ],
        synced_at=datetime(2026, 8, 22, 11, 0),
    )

    assert result == (1, 1, 1)
    updated = (
        db.execute(models.CellarDeck.__table__.select().where(models.CellarDeck.source_key == "gsheet:old:1"))
        .mappings()
        .one()
    )
    assert updated["name"] == "New name"
    assert updated["decklist_url"] == "https://example.test/new"
    assert updated["decklist_updated_on"] == date(2026, 7, 1)
    assert updated["available"] is False
    assert (
        db.execute(models.CellarDeck.__table__.select().where(models.CellarDeck.source_key == "gsheet:removed:1"))
        .mappings()
        .one()["active"]
        is False
    )


def test_catalog_refresh_uses_persisted_fifteen_minute_ttl(db):
    class Source:
        def __init__(self):
            self.calls = 0

        def fetch(self):
            self.calls += 1
            return [CatalogEntry("gsheet:deck:1", "Deck", "Deck")]

    source = Source()
    service = CellarService(db)
    now = datetime(2026, 8, 22, 10, 0)

    assert service.refresh_catalog_from_sheet(now=now, source=source) == (1, 0, 0)
    assert service.refresh_catalog_from_sheet(now=now + timedelta(minutes=14), source=source) is None
    assert service.refresh_catalog_from_sheet(
        now=now + CELLAR_CATALOG_REFRESH_INTERVAL,
        source=source,
    ) == (0, 1, 0)
    assert source.calls == 2


def test_unavailable_sheet_deck_cannot_be_reserved(db, user_svc):
    service = CellarService(db)
    service.sync_catalog([CatalogEntry("gsheet:deck:1", "Deck", "Deck", notes="занято", available=False)])
    deck = service.catalog(date(2026, 8, 24))[0]
    user = user_svc.get_or_create(tg_id=1001, first_name="Alice")

    with pytest.raises(CellarDeckUnavailable, match="недоступна"):
        service.reserve(deck_id=deck.id, user_id=user.id, event_date=date(2026, 8, 24), today=date(2026, 8, 24))
