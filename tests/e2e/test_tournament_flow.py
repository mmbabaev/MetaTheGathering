"""E2E regression tests: create → import → export flow."""

import io
from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest
from sqlalchemy import select

from core.models import Participant, User
from core.schemas import TournamentCreate
from services.aetherhub_import_service import AetherhubImportService
from services.aetherhub_models import AetherhubPairing, AetherhubRound, AetherhubTournamentData
from services.aetherhub_service import AetherhubService
from services.export import ExportService
from services.tournament import TournamentService
from services.user import UserService
from tests.e2e.conftest import CHAT_ID

_FIXTURES = Path(__file__).resolve().parents[2] / "scripts" / "aetherhub" / "fixtures"
_TID = "99291"
_URL = f"https://aetherhub.com/Tourney/RoundTourney/{_TID}"


def _scraper_99291():
    """Mock scraper backed by saved 99291 HTML fixtures."""
    url_map = {
        _URL: (_FIXTURES / "99291_main.html").read_text(),
        **{
            f"https://aetherhub.com/Tourney/RoundTourneyPublicPairings?id={_TID}&p={r}": (
                _FIXTURES / f"99291_pairings_p{r}.html"
            ).read_text()
            for r in range(1, 5)
        },
    }
    scraper = MagicMock()
    # main page ?p=N not in the saved fixtures → empty → fetch_tournament falls
    # back to the public pairings endpoint (the saved fixtures).
    scraper.get.side_effect = lambda url, **_: MagicMock(text=url_map.get(url, "<html><body></body></html>"))
    return scraper


def _fixtures_available():
    return all(
        (_FIXTURES / f).exists() for f in ["99291_main.html", *[f"99291_pairings_p{r}.html" for r in range(1, 5)]]
    )


def test_create_tournament(svc):
    t = svc.create_tournament(TournamentCreate(title="Test", chat_id=CHAT_ID))
    assert t.id is not None
    assert t.title == "Test"
    assert t.status.value == "registration"


def test_delete_last(db, svc):
    t1 = svc.create_tournament(TournamentCreate(title="First", chat_id=CHAT_ID))
    svc.close_tournament(t1.id)
    t2 = svc.create_tournament(TournamentCreate(title="Second", chat_id=CHAT_ID))
    svc.close_tournament(t2.id)

    last = svc.list_tournaments_for_chat(CHAT_ID, limit=1)
    assert last[0].title == "Second"

    svc.delete_tournament(last[0].id)
    remaining = svc.list_tournaments_for_chat(CHAT_ID, limit=10)
    assert len(remaining) == 1
    assert remaining[0].title == "First"


def test_import_aetherhub(db, tournament, aetherhub_data):
    result = AetherhubImportService(db).import_tournament(tournament.id, aetherhub_data)

    assert result.registered == 4
    assert result.already_registered == 0
    assert result.pairings_saved == 8  # 4 записи × 2 раунда

    participants = TournamentService(db).list_participants_for_tournament(tournament.id)
    assert len(participants) == 4

    rows = db.execute(
        select(Participant, User)
        .join(User, User.id == Participant.user_id)
        .where(Participant.tournament_id == tournament.id)
    ).all()
    places = {u.first_name: p.final_place for p, u in rows}
    assert places["Иван"] == 1
    assert places["Сидор"] == 2


def test_import_idempotent(db, tournament, aetherhub_data):
    svc = AetherhubImportService(db)
    svc.import_tournament(tournament.id, aetherhub_data)
    result2 = svc.import_tournament(tournament.id, aetherhub_data)

    assert result2.registered == 0
    assert result2.already_registered == 4


def test_export_excel(db, tournament, aetherhub_data, tmp_path):
    AetherhubImportService(db).import_tournament(tournament.id, aetherhub_data)
    data, filename = ExportService(db).export_participants_excel(tournament.id)

    assert len(data) > 0
    assert filename.endswith(".xlsx")
    out = tmp_path / filename
    out.write_bytes(data)
    assert out.exists()


def test_full_flow(db, svc, aetherhub_data, tmp_path):
    """Полный флоу: создать → импорт → экспорт → удалить."""
    t = svc.create_tournament(TournamentCreate(title="Pauper Friday #99", chat_id=CHAT_ID))

    result = AetherhubImportService(db).import_tournament(t.id, aetherhub_data)
    assert result.registered == 4

    data, filename = ExportService(db).export_participants_excel(t.id)
    assert len(data) > 0
    (tmp_path / filename).write_bytes(data)

    svc.close_tournament(t.id)
    last = svc.list_tournaments_for_chat(CHAT_ID, limit=1)
    assert last[0].id == t.id
    svc.delete_tournament(last[0].id)
    assert svc.list_tournaments_for_chat(CHAT_ID) == []


# ── get_player_opponents: e2e ─────────────────────────────────────────────────


def test_get_player_opponents_full_flow(db, svc, aetherhub_data):
    """Полный флоу: create → import → get_player_opponents для конкретного участника."""
    t = svc.create_tournament(TournamentCreate(title="Opponents Flow", chat_id=CHAT_ID))
    AetherhubImportService(db).import_tournament(t.id, aetherhub_data)

    import_svc = AetherhubImportService(db)
    ivan = import_svc.find_user_by_name("Иван Иванов")
    assert ivan is not None, "Иван Иванов not found after import"

    ivan_p = import_svc._get_participant(t.id, ivan.id)
    assert ivan_p is not None

    opps, err = import_svc.get_player_opponents(t.id, ivan_p.id)
    assert err is None
    assert len(opps) == 2

    round_to_opponent = {o.round_number: o.opponent_name for o in opps}
    assert round_to_opponent[1] == "Пётр Петров"
    assert round_to_opponent[2] == "Сидор Сидоров"

    assert all(o.opponent_user is not None for o in opps)
    assert all(o.opponent_participant is not None for o in opps)

    opp_names_in_db = {o.opponent_user.first_name for o in opps}
    assert "Пётр" in opp_names_in_db
    assert "Сидор" in opp_names_in_db


def test_get_player_opponents_bye(db, svc):
    """Bye отображается как opponent_name=None."""
    t = svc.create_tournament(TournamentCreate(title="Bye Test", chat_id=CHAT_ID))
    data = AetherhubTournamentData(
        url="http://x",
        players=["Иван Иванов"],
        rounds=[AetherhubRound(number=1, pairings=[AetherhubPairing(player="Иван Иванов", opponent=None)])],
        standings=["Иван Иванов"],
    )
    import_svc = AetherhubImportService(db)
    import_svc.import_tournament(t.id, data)
    ivan = import_svc.find_user_by_name("Иван Иванов")
    ivan_p = import_svc._get_participant(t.id, ivan.id)

    opps, err = import_svc.get_player_opponents(t.id, ivan_p.id)
    assert err is None
    assert len(opps) == 1
    assert opps[0].opponent_name is None
    assert opps[0].opponent_user is None


# ── CLI-style integration: реальные HTML fixtures, полный pipeline ────────────


@pytest.mark.skipif(not _fixtures_available(), reason="99291 HTML fixtures missing")
def test_cli_import_export_sorted_by_final_standings(db, svc):
    """Повторяет то, что делает CLI: create → fetch → import → export-excel.

    Использует сохранённые HTML fixtures турнира 99291 (39 игроков, 4 раунда).
    Проверяет что выгрузка отсортирована по финальным стендингам:
    Федулов Ринат должен быть первым, Нагорнов Владимир — последним.
    """
    # 1. Создаём турнир (cli.py tournament create)
    t = svc.create_tournament(TournamentCreate(title="Pauper Friday #99291", chat_id=CHAT_ID))

    # 2. Парсим данные с aetherhub (cli.py tournament import <url>)
    #    HTTP-запросы заменены на локальные HTML fixtures
    fetch_svc = AetherhubService(scraper=_scraper_99291())
    data = fetch_svc.fetch_tournament(_URL)

    assert len(data.players) == 39
    assert len(data.rounds) == 4
    assert len(data.standings) == 39
    assert data.standings[0] == "Федулов Ринат"
    assert data.standings[-1] == "Нагорнов Владимир"

    # 3. Импортируем в БД
    result = AetherhubImportService(db).import_tournament(t.id, data)
    assert result.registered == 39
    assert result.already_registered == 0

    # 4. Экспортируем в Excel (cli.py tournament export-excel)
    excel_bytes, filename = ExportService(db).export_participants_excel(t.id)
    assert filename.endswith(".xlsx")

    # 5. Проверяем порядок строк в Excel — должен совпадать с финальными стендингами
    ws = openpyxl.load_workbook(io.BytesIO(excel_bytes)).active
    names = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]

    assert len(names) == 39
    # Победитель первым, последний место в конце
    assert names[0] == "Ринат Федулов"
    assert names[-1] == "Владимир Нагорнов"
    # Топ-3 совпадает со стендингами (стендинги: Федулов, Рябинин, Юдин)
    assert names[1] == "Андрей Рябинин"
    assert names[2] == "Антон Юдин"
