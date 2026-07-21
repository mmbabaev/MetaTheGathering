"""Part 3: export methods sort participants by final_place (nulls last)."""

import csv
import io

import openpyxl
import pytest

from core import models
from services.export import ExportService


def _add_participant(db, tournament, user, final_place=None, archetype=None):
    p = models.Participant(
        tournament_id=tournament.id,
        user_id=user.id,
        final_place=final_place,
        archetype_id=archetype.id if archetype else None,
    )
    db.add(p)
    db.commit()
    return p


@pytest.fixture
def export_svc(db):
    return ExportService(db)


@pytest.fixture
def user_carol(user_svc):
    return user_svc.get_or_create(tg_id=1003, username="carol", first_name="Carol")


# ── TestExportPlayersListSorting ──────────────────────────────────────────────


class TestExportPlayersListSorting:
    def test_sorted_by_final_place(self, export_svc, db, tournament, user_alice, user_bob, user_carol):
        _add_participant(db, tournament, user_carol, final_place=1)
        _add_participant(db, tournament, user_alice, final_place=2)
        _add_participant(db, tournament, user_bob, final_place=3)
        result = export_svc.export_players_list(tournament.id)
        names = result.splitlines()
        assert names == ["Carol", "Alice", "Bob"]

    def test_participants_without_place_go_last(self, export_svc, db, tournament, user_alice, user_bob):
        _add_participant(db, tournament, user_bob, final_place=1)
        _add_participant(db, tournament, user_alice, final_place=None)
        result = export_svc.export_players_list(tournament.id)
        names = result.splitlines()
        assert names[0] == "Bob"
        assert names[1] == "Alice"

    def test_nulls_sorted_alphabetically_among_themselves(
        self, export_svc, db, tournament, user_alice, user_bob, user_carol
    ):
        _add_participant(db, tournament, user_carol, final_place=None)
        _add_participant(db, tournament, user_alice, final_place=None)
        _add_participant(db, tournament, user_bob, final_place=None)
        result = export_svc.export_players_list(tournament.id)
        names = result.splitlines()
        assert names == sorted(names)


# ── TestExportExcelSorting ────────────────────────────────────────────────────


class TestExportExcelSorting:
    def _read_names(self, excel_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        return [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]

    def test_rows_sorted_by_final_place(self, export_svc, db, tournament, user_alice, user_bob, user_carol):
        _add_participant(db, tournament, user_carol, final_place=1)
        _add_participant(db, tournament, user_alice, final_place=2)
        _add_participant(db, tournament, user_bob, final_place=3)
        data, _ = export_svc.export_participants_excel(tournament.id)
        names = self._read_names(data)
        assert names == ["Carol", "Alice", "Bob"]

    def test_nulls_at_end_of_excel(self, export_svc, db, tournament, user_alice, user_bob):
        _add_participant(db, tournament, user_bob, final_place=1)
        _add_participant(db, tournament, user_alice, final_place=None)
        data, _ = export_svc.export_participants_excel(tournament.id)
        names = self._read_names(data)
        assert names[0] == "Bob"
        assert names[1] == "Alice"

    def test_general_type_column(self, export_svc, db, tournament, user_alice, arch_svc):
        tournament_row = db.query(models.Tournament).filter_by(id=tournament.id).one()
        tournament_row.decks_hidden = False  # колонки колод показываются только когда раскрыты
        db.commit()
        arch = arch_svc.get_or_create_by_name("Blue Delver")  # общий тип → «Blue Terror»
        _add_participant(db, tournament, user_alice, final_place=1, archetype=arch)
        data, _ = export_svc.export_participants_excel(tournament.id)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.cell(row=1, column=4).value == "Колода"
        assert ws.cell(row=1, column=5).value == "Общий тип"
        assert ws.cell(row=2, column=4).value == "Blue Delver"
        assert ws.cell(row=2, column=5).value == "Blue Terror"


# ── TestExportCsvSorting ──────────────────────────────────────────────────────


class TestExportCsvSorting:
    def _parse_csv_usernames(self, csv_text):
        reader = csv.DictReader(io.StringIO(csv_text))
        return [row["username"] for row in reader]

    def test_csv_sorted_by_final_place(self, export_svc, db, tournament, user_alice, user_bob, user_carol):
        _add_participant(db, tournament, user_carol, final_place=1)
        _add_participant(db, tournament, user_alice, final_place=2)
        _add_participant(db, tournament, user_bob, final_place=3)
        csv_text = export_svc.export_participants_csv(tournament.id)
        usernames = self._parse_csv_usernames(csv_text)
        assert usernames == ["carol", "alice", "bob"]

    def test_csv_nulls_last(self, export_svc, db, tournament, user_alice, user_bob):
        _add_participant(db, tournament, user_bob, final_place=1)
        _add_participant(db, tournament, user_alice, final_place=None)
        csv_text = export_svc.export_participants_csv(tournament.id)
        usernames = self._parse_csv_usernames(csv_text)
        assert usernames[0] == "bob"
        assert usernames[1] == "alice"
