"""Tests for decks_hidden feature: spoiler wrapping, reveal button, service method."""

import io
from unittest.mock import patch

import openpyxl
import pytest

from bot.handlers.admin import AdminHandler
from bot.handlers.player import PlayerHandler
from bot.keyboards import (
    CB_HIDE_DECKS,
    CB_REVEAL_DECKS,
    CB_REVEAL_DECKS_CANCEL,
    CB_REVEAL_DECKS_CONFIRM,
    admin_more_keyboard,
    reveal_decks_confirm_keyboard,
)
from bot.messages import DECKS_REVEALED, format_tournament_status
from core import models
from core.schemas import TournamentCreate
from services.export import ExportService
from services.tournament import TournamentService

CHAT_ID = 777


@pytest.fixture
def tournament(svc):
    return svc.create_tournament(TournamentCreate(title="Hidden Cup", chat_id=CHAT_ID, slug="hidden"))


@pytest.fixture
def handler(svc, user_svc, arch_svc, keyboards, aetherhub_svc, features):
    return PlayerHandler(svc, user_svc, arch_svc, keyboards, aetherhub_svc, features)


@pytest.fixture
def admin_handler(svc, user_svc, arch_svc, keyboards, features):
    return AdminHandler(svc, user_svc, arch_svc, keyboards, features)


# ===== TournamentService.set_decks_hidden =====


class TestSetDecksHidden:
    def test_default_is_true(self, tournament):
        assert tournament.decks_hidden is True

    def test_set_false(self, svc, tournament):
        updated = svc.set_decks_hidden(tournament.id, hidden=False)
        assert updated.decks_hidden is False

    def test_set_true_again(self, svc, tournament):
        svc.set_decks_hidden(tournament.id, hidden=False)
        updated = svc.set_decks_hidden(tournament.id, hidden=True)
        assert updated.decks_hidden is True


# ===== format_tournament_status with decks_hidden =====


class TestFormatTournamentStatus:
    def test_hidden_shows_placeholder(self, db, svc, user_svc, arch_svc, tournament, archetype_burn):
        user = user_svc.get_or_create(tg_id=4001, username="u", first_name="Ivan")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)
        participants = svc.list_participants_for_tournament(tournament.id)

        text = format_tournament_status("T", "Регистрация", participants, decks_hidden=True)
        assert "▓▓▓" in text
        assert "Burn" not in text

    def test_not_hidden_shows_plain_archetype(self, db, svc, user_svc, arch_svc, tournament, archetype_burn):
        user = user_svc.get_or_create(tg_id=4002, username="u2", first_name="Maria")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)
        participants = svc.list_participants_for_tournament(tournament.id)

        text = format_tournament_status("T", "Регистрация", participants, decks_hidden=False)
        assert "▓▓▓" not in text
        assert "Burn" in text

    def test_no_archetype_never_has_placeholder(self, db, svc, user_svc, tournament):
        user = user_svc.get_or_create(tg_id=4003, username="u3", first_name="Oleg")
        db.add(models.Participant(tournament_id=tournament.id, user_id=user.id))
        db.commit()
        participants = svc.list_participants_for_tournament(tournament.id)

        text = format_tournament_status("T", "Регистрация", participants, decks_hidden=True)
        assert "▓▓▓" not in text
        assert "не указана" in text


# ===== PlayerHandler.handle_tournament_public_status =====


class TestHandleTournamentPublicStatus:
    def test_decks_hidden_shows_placeholder(self, handler, svc, user_svc, arch_svc, tournament, archetype_burn):
        user = user_svc.get_or_create(tg_id=5001, username="p", first_name="Player")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        result = handler.handle_tournament_public_status(tournament.id)
        assert "▓▓▓" in result.text
        assert "Burn" not in result.text

    def test_decks_revealed_shows_name(self, handler, svc, user_svc, arch_svc, tournament, archetype_burn):
        svc.set_decks_hidden(tournament.id, hidden=False)
        user = user_svc.get_or_create(tg_id=5002, username="p2", first_name="Player2")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        result = handler.handle_tournament_public_status(tournament.id)
        assert "▓▓▓" not in result.text
        assert "Burn" in result.text


# ===== Tournament card keyboard: reveal button =====


class TestTournamentCardKeyboard:
    def test_reveal_button_not_in_tournament_card(self, handler, user_svc, svc, arch_svc, tournament, archetype_burn):
        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = [9001]
            result = handler.handle_tournaments(tg_id=9001)

        buttons_flat = [btn for row in result.keyboard.inline_keyboard for btn in row]
        assert not any(b.callback_data.startswith(CB_REVEAL_DECKS) for b in buttons_flat)

    def test_non_admin_never_sees_reveal_button(self, handler, tournament):
        result = handler.handle_tournaments(tg_id=99999)
        buttons_flat = [btn for row in result.keyboard.inline_keyboard for btn in row]
        assert not any(b.callback_data.startswith(CB_REVEAL_DECKS) for b in buttons_flat)


# ===== Admin view also respects decks_hidden =====


class TestAdminViewDecksHidden:
    def test_admin_status_hides_decks_when_hidden(self, admin_handler, svc, user_svc, tournament, archetype_burn):
        user = user_svc.get_or_create(tg_id=7001, username="p", first_name="Player")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = [8888]
            result = admin_handler.handle_admin_status(tg_id=8888, tournament_id=tournament.id)

        assert "▓▓▓" in result.text
        assert "Burn" not in result.text

    def test_admin_status_shows_decks_after_reveal(self, admin_handler, svc, user_svc, tournament, archetype_burn):
        svc.set_decks_hidden(tournament.id, hidden=False)
        user = user_svc.get_or_create(tg_id=7002, username="p2", first_name="Player2")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = [8888]
            result = admin_handler.handle_admin_status(tg_id=8888, tournament_id=tournament.id)

        assert "▓▓▓" not in result.text
        assert "Burn" in result.text


# ===== AdminHandler.handle_reveal_decks =====


class TestHandleRevealDecks:
    def test_reveals_decks_and_returns_status(self, admin_handler, svc, user_svc, tournament, archetype_burn):
        user = user_svc.get_or_create(tg_id=6001, username="p", first_name="Player")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = [8001]
            result = admin_handler.handle_reveal_decks(tg_id=8001, tournament_id=tournament.id)

        assert not result.is_alert
        assert DECKS_REVEALED in result.text
        updated = svc.get_active_tournament_for_chat(CHAT_ID)
        assert updated.decks_hidden is False

    def test_non_admin_returns_alert(self, admin_handler, tournament):
        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = []
            result = admin_handler.handle_reveal_decks(tg_id=99999, tournament_id=tournament.id)

        assert result.is_alert

    def test_tournament_not_found_returns_alert(self, admin_handler):
        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = [8001]
            result = admin_handler.handle_reveal_decks(tg_id=8001, tournament_id=99999)

        assert result.is_alert


# ===== AdminHandler.handle_hide_decks =====


class TestHandleHideDecks:
    def test_hides_decks_and_returns_status(self, admin_handler, svc, user_svc, tournament, archetype_burn):
        svc.set_decks_hidden(tournament.id, hidden=False)
        user = user_svc.get_or_create(tg_id=6010, username="p", first_name="Player")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = [8001]
            result = admin_handler.handle_hide_decks(tg_id=8001, tournament_id=tournament.id)

        assert not result.is_alert
        assert "скрыты" in result.text.lower()
        assert svc.get_active_tournament_for_chat(CHAT_ID).decks_hidden is True

    def test_non_admin_returns_alert(self, admin_handler, tournament):
        with patch("services.user.settings") as mock_settings:
            mock_settings.admin_ids = []
            result = admin_handler.handle_hide_decks(tg_id=99999, tournament_id=tournament.id)

        assert result.is_alert


# ===== reveal_decks_confirm_keyboard =====


class TestRevealDecksConfirmKeyboard:
    def test_has_confirm_and_cancel(self):
        kb = reveal_decks_confirm_keyboard(tournament_id=42)
        buttons = {b.callback_data for row in kb.inline_keyboard for b in row}
        assert f"{CB_REVEAL_DECKS_CONFIRM}:42" in buttons
        assert f"{CB_REVEAL_DECKS_CANCEL}:42" in buttons


# ===== admin_more_keyboard show/hide decks button =====


class TestAdminMoreKeyboardDecks:
    def test_shows_reveal_when_hidden(self):
        kb = admin_more_keyboard(tournament_id=1, decks_hidden=True)
        cbs = [b.callback_data for row in kb.inline_keyboard for b in row]
        assert any(cb.startswith(CB_REVEAL_DECKS + ":") for cb in cbs)

    def test_shows_hide_when_revealed(self):
        kb = admin_more_keyboard(tournament_id=1, decks_hidden=False)
        cbs = [b.callback_data for row in kb.inline_keyboard for b in row]
        assert any(cb.startswith(CB_HIDE_DECKS + ":") for cb in cbs)


# ===== Excel export respects decks_hidden =====


class TestExportExcelDecksHidden:
    def test_deck_column_absent_when_hidden(self, svc, user_svc, arch_svc, tournament, archetype_burn):
        user = user_svc.get_or_create(tg_id=7001, username="u", first_name="Test")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        data, _ = ExportService(svc.db).export_participants_excel(tournament.id)
        ws = openpyxl.load_workbook(io.BytesIO(data)).active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Колода" not in headers

    def test_deck_column_present_when_revealed(self, svc, user_svc, arch_svc, tournament, archetype_burn):
        svc.set_decks_hidden(tournament.id, hidden=False)
        user = user_svc.get_or_create(tg_id=7002, username="u", first_name="Test")
        svc.register_participant(tournament_id=tournament.id, user_id=user.id, archetype_id=archetype_burn.id)

        data, _ = ExportService(svc.db).export_participants_excel(tournament.id)
        ws = openpyxl.load_workbook(io.BytesIO(data)).active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        assert "Колода" in headers
